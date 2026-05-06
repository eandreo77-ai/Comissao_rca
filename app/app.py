"""
Sistema de Importação de Comissões RCA
Interface Streamlit — visual padrão ROFE
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import streamlit as st
import pandas as pd
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import ORACLE_CONFIG, CODCONTA_PADRAO, CODFILIAL_PADRAO, TIPOSERVICO, ROTINA_749
from oracle_db import OracleConnection, GravadorPCLANC, testar_conexao
from excel_parser import ler_excel
import audit

# === ROFE_VISUAL_PATCH ===
from styles import aplicar_visual, header, secao
import auth
auth.gate_login()  # bloqueia execução se não logado
_u = auth.usuario_logado()
aplicar_visual()
header(
    titulo='Importacao de Comissao RCA',
    subtitulo='Lancamento no contas a pagar (rotina 749)',
    icone='\U0001f4b0',
    usuario=_u['email'] if _u else None,
)
# === FIM_ROFE_VISUAL_PATCH ===


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
def brl(valor: float) -> str:
    return "R$ " + f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def badge(texto: str, cor: str) -> str:
    cores = {
        "azul":    "#2563eb",
        "verde":   "#16a34a",
        "vermelho":"#dc2626",
        "cinza":   "#64748b",
        "navy":    "#0f172a",
    }
    bg = cores.get(cor, cor)
    return (
        f'<span style="display:inline-block;padding:5px 14px;border-radius:20px;'
        f'background:{bg};color:#fff;font-size:13px;font-weight:600;margin:3px 3px 3px 0;">'
        f'{texto}</span>'
    )


def gerar_modelo_excel() -> bytes:
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.worksheet.datavalidation import DataValidation

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Comissoes"

    # Colunas: apenas as usadas pelo parser
    # Fixas obrigatórias / opcionais + duas datas de exemplo
    fixed = [
        ("parceiro(COD)", 16),   # A — CODUSUR (obrigatório)
        ("RCA",           28),   # B — nome referência (opcional)
        ("contadebito",   14),   # C — conta contábil (obrigatório)
        ("historico",     32),   # D — descrição (opcional)
    ]
    datas = [("13/04/2026", 16), ("30/04/2026", 16)]   # E, F
    cols  = fixed + datas
    n_fixed     = len(fixed)
    n_rows_data = 1
    max_row     = 1 + n_rows_data   # linha 2

    # ── Estilos ──────────────────────────────────────────────────────
    sd   = Side(style="thin", color="AAAAAA")
    brd  = Border(left=sd, right=sd, top=sd, bottom=sd)
    ct   = Alignment(horizontal="center", vertical="center")
    hf   = Font(bold=True, color="FFFFFF", size=11)
    hbl  = PatternFill("solid", fgColor="0F172A")   # cabeçalho — navy
    ferr = PatternFill("solid", fgColor="FEE2E2")   # CF erro — vermelho

    # ── Cabeçalho ────────────────────────────────────────────────────
    for ci, (h, w) in enumerate(cols, 1):
        c = ws.cell(1, ci, h)
        c.font = hf
        c.fill = hbl
        c.border = brd
        c.alignment = ct
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    # ── 1 linha de exemplo — sem cor ─────────────────────────────────
    exemplo = [11, "TELMI TEIXEIRA DO LAGO", 100010, "COMISSAO ABR/2026", 7219.48, 7219.48]
    for ci, val in enumerate(exemplo, 1):
        is_p = ci > n_fixed
        c = ws.cell(2, ci, val)
        c.border = brd
        c.alignment = Alignment(vertical="center")
        if ci == 1:
            c.number_format = "0"
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif ci == 3:
            c.number_format = "0"
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif is_p and val is not None:
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Validação de dados ───────────────────────────────────────────
    # parceiro(COD): inteiro entre 1 e 999999
    dv_cod = DataValidation(
        type="whole", operator="between", formula1="1", formula2="999999",
        error="Informe um codigo RCA inteiro valido (ex: 11)",
        errorTitle="CODUSUR invalido", showErrorMessage=True,
    )
    dv_cod.sqref = f"A2:A{max_row + 50}"
    ws.add_data_validation(dv_cod)

    # contadebito: inteiro > 0
    dv_conta = DataValidation(
        type="whole", operator="greaterThan", formula1="0",
        error="Informe o codigo da conta contabil (ex: 100010)",
        errorTitle="Conta invalida", showErrorMessage=True,
    )
    dv_conta.sqref = f"C2:C{max_row + 50}"
    ws.add_data_validation(dv_conta)

    # colunas de parcela: decimal > 0
    for col_idx in range(n_fixed + 1, len(cols) + 1):
        col_letter = get_column_letter(col_idx)
        dv_val = DataValidation(
            type="decimal", operator="greaterThan", formula1="0",
            error="Informe o valor da parcela maior que zero",
            errorTitle="Valor invalido", showErrorMessage=True,
        )
        dv_val.sqref = f"{col_letter}2:{col_letter}{max_row + 50}"
        ws.add_data_validation(dv_val)

    # ── Formatação condicional — todos obrigatórios ──────────────────
    limite = max_row + 50

    # A — parceiro(COD): vazio ou não-numérico → vermelho
    ws.conditional_formatting.add(
        f"A2:A{limite}",
        FormulaRule(formula=["OR(ISBLANK(A2),NOT(ISNUMBER(A2)))"],
                    fill=ferr, stopIfTrue=True)
    )
    # B — RCA: vazio quando linha tem dado → vermelho
    ws.conditional_formatting.add(
        f"B2:B{limite}",
        FormulaRule(formula=["AND(NOT(ISBLANK(A2)),ISBLANK(B2))"],
                    fill=ferr, stopIfTrue=True)
    )
    # C — contadebito: vazio ou não-numérico → vermelho
    ws.conditional_formatting.add(
        f"C2:C{limite}",
        FormulaRule(formula=["OR(ISBLANK(C2),NOT(ISNUMBER(C2)))"],
                    fill=ferr, stopIfTrue=True)
    )
    # D — historico: vazio quando linha tem dado → vermelho
    ws.conditional_formatting.add(
        f"D2:D{limite}",
        FormulaRule(formula=["AND(NOT(ISBLANK(A2)),ISBLANK(D2))"],
                    fill=ferr, stopIfTrue=True)
    )
    # Colunas de parcela: valor negativo ou zero → vermelho
    for col_idx in range(n_fixed + 1, len(cols) + 1):
        col_letter = get_column_letter(col_idx)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{limite}",
            FormulaRule(formula=[f"AND(NOT(ISBLANK({col_letter}2)),{col_letter}2<=0)"],
                        fill=ferr, stopIfTrue=True)
        )

    # ── Aba Instruções ───────────────────────────────────────────────
    wi = wb.create_sheet("Instrucoes")
    wi.column_dimensions["A"].width = 18
    wi.column_dimensions["B"].width = 55
    wi.column_dimensions["C"].width = 14
    wi.column_dimensions["D"].width = 30

    for cc in ["A1","B1","C1","D1"]:
        wi[cc].font = hf; wi[cc].fill = hbl; wi[cc].border = brd
        wi[cc].alignment = ct
    wi["A1"] = "Coluna"; wi["B1"] = "Descricao"
    wi["C1"] = "Obrigatorio"; wi["D1"] = "Alerta / Validacao"

    ins = [
        ("parceiro(COD)", "Codigo do RCA no WinThor (inteiro)",          "SIM", "Vermelho se vazio ou nao-numerico"),
        ("RCA",           "Nome do RCA (referencia visual)",             "SIM", "Vermelho se linha preenchida sem RCA"),
        ("contadebito",   "Codigo da conta contabil (ex: 100010)",       "SIM", "Vermelho se vazio ou nao-numerico"),
        ("historico",     "Descricao do lancamento (COMISSAO ABR/2026)", "SIM", "Vermelho se linha preenchida sem historico"),
        ("DD/MM/AAAA",    "Cabecalho = data de vencimento da parcela",   "SIM", "Vermelho se valor <= 0"),
        ("DD/MM/AAAA ...", "Adicione quantas colunas de data precisar",  "SIM", "Vermelho se valor <= 0"),
    ]
    bf = Font(bold=True)
    for ri2, (campo, desc, ob, alerta) in enumerate(ins, 2):
        wi.cell(ri2, 1, campo).font = bf
        wi.cell(ri2, 2, desc)
        wi.cell(ri2, 3, ob).alignment = Alignment(horizontal="center")
        wi.cell(ri2, 4, alerta)
        for cc in range(1, 5):
            wi.cell(ri2, cc).border = brd
        wi.row_dimensions[ri2].height = 18

    nr = len(ins) + 3
    c_regra = wi.cell(nr, 1, "REGRA:")
    c_regra.font = Font(bold=True, color="C00000")
    wi.cell(nr, 2, "Colunas com cabecalho no formato DD/MM/AAAA viram lancamentos separados na PCLANC.")

    buf = BytesIO(); wb.save(buf); return buf.getvalue()


# ─────────────────────────────────────────────
# Página
# ─────────────────────────────────────────────
# REMOVED_BY_VISUAL_PATCH: st.set_page_config(page_title="ROFE | Comissão RCA", page_icon="💰", layout="wide")

# # REMOVED_BY_DARK_THEME_PATCH: st.markdown("""
# # REMOVED_BY_DARK_THEME_PATCH: <style>
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* ===== FONT ===== */
# # REMOVED_BY_DARK_THEME_PATCH: @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: html, body, [class*="css"] {
# # REMOVED_BY_DARK_THEME_PATCH:     font-family: 'Inter', sans-serif;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* ===== SIDEBAR — fundo e catch-all de texto ===== */
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"],
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] > div,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] [data-testid="stVerticalBlock"],
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] .stMarkdown,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] .element-container,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] .stSelectbox,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] .stTextInput {
# # REMOVED_BY_DARK_THEME_PATCH:     background-color: #0f172a !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* Texto padrão — tudo branco/cinza claro */
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] *,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] p,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] span,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] small,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] div,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] label,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] h1,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] h2,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] h3,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] li {
# # REMOVED_BY_DARK_THEME_PATCH:     color: #cbd5e1 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     background-color: transparent !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* Títulos um pouco mais brilhantes */
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] h1,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] h2,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] h3 {
# # REMOVED_BY_DARK_THEME_PATCH:     color: #f8fafc !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] hr {
# # REMOVED_BY_DARK_THEME_PATCH:     border-color: #1e293b !important;
# # REMOVED_BY_DARK_THEME_PATCH:     margin: 12px 0 !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* === Navegação entre páginas (sidebar) === */
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a {
# # REMOVED_BY_DARK_THEME_PATCH:     border-radius: 8px !important;
# # REMOVED_BY_DARK_THEME_PATCH:     margin: 2px 6px !important;
# # REMOVED_BY_DARK_THEME_PATCH:     padding: 10px 12px !important;
# # REMOVED_BY_DARK_THEME_PATCH:     font-weight: 500 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     transition: all .15s ease-in-out;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a:hover {
# # REMOVED_BY_DARK_THEME_PATCH:     background-color: #1e293b !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: /* Página selecionada */
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a[aria-current="page"] {
# # REMOVED_BY_DARK_THEME_PATCH:     background: linear-gradient(90deg, #2563eb 0%, #1d4ed8 100%) !important;
# # REMOVED_BY_DARK_THEME_PATCH:     border-left: 4px solid #ef4444 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     box-shadow: 0 2px 8px rgba(37,99,235,.4) !important;
# # REMOVED_BY_DARK_THEME_PATCH:     font-weight: 700 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     padding-left: 14px !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a[aria-current="page"] *,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a[aria-current="page"] span {
# # REMOVED_BY_DARK_THEME_PATCH:     color: #ffffff !important;
# # REMOVED_BY_DARK_THEME_PATCH:     font-weight: 700 !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* ===== INPUTS sidebar — padrão único todos os estados ===== */
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] input,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] select {
# # REMOVED_BY_DARK_THEME_PATCH:     background-color: #1e293b !important;
# # REMOVED_BY_DARK_THEME_PATCH:     color: #f8fafc !important;
# # REMOVED_BY_DARK_THEME_PATCH:     -webkit-text-fill-color: #f8fafc !important;
# # REMOVED_BY_DARK_THEME_PATCH:     border: 1px solid #334155 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     border-radius: 8px !important;
# # REMOVED_BY_DARK_THEME_PATCH:     opacity: 1 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     caret-color: transparent !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* Disabled */
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] input:disabled,
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] select:disabled {
# # REMOVED_BY_DARK_THEME_PATCH:     background-color: #1e293b !important;
# # REMOVED_BY_DARK_THEME_PATCH:     color: #94a3b8 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     -webkit-text-fill-color: #94a3b8 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     border: 1px solid #1e293b !important;
# # REMOVED_BY_DARK_THEME_PATCH:     opacity: 1 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     cursor: default !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] input:focus {
# # REMOVED_BY_DARK_THEME_PATCH:     border: 1px solid #3b82f6 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     caret-color: #f8fafc !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* Dropdown do selectbox */
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] [data-testid="stSelectbox"] > div > div {
# # REMOVED_BY_DARK_THEME_PATCH:     background-color: #1e293b !important;
# # REMOVED_BY_DARK_THEME_PATCH:     border: 1px solid #334155 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     border-radius: 8px !important;
# # REMOVED_BY_DARK_THEME_PATCH:     color: #f8fafc !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* Seta do selectbox */
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] [data-testid="stSelectbox"] svg {
# # REMOVED_BY_DARK_THEME_PATCH:     fill: #94a3b8 !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* Botões sidebar */
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] button {
# # REMOVED_BY_DARK_THEME_PATCH:     background-color: #22c55e !important;
# # REMOVED_BY_DARK_THEME_PATCH:     color: #ffffff !important;
# # REMOVED_BY_DARK_THEME_PATCH:     -webkit-text-fill-color: #ffffff !important;
# # REMOVED_BY_DARK_THEME_PATCH:     border-radius: 8px !important;
# # REMOVED_BY_DARK_THEME_PATCH:     font-weight: 600 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     border: none !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: section[data-testid="stSidebar"] button:hover {
# # REMOVED_BY_DARK_THEME_PATCH:     background-color: #16a34a !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* ===== HEADER ===== */
# # REMOVED_BY_DARK_THEME_PATCH: .erp-header {
# # REMOVED_BY_DARK_THEME_PATCH:     background: #0f172a;
# # REMOVED_BY_DARK_THEME_PATCH:     padding: 20px 24px;
# # REMOVED_BY_DARK_THEME_PATCH:     border-radius: 12px;
# # REMOVED_BY_DARK_THEME_PATCH:     margin-bottom: 20px;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: .erp-title {
# # REMOVED_BY_DARK_THEME_PATCH:     color: white;
# # REMOVED_BY_DARK_THEME_PATCH:     font-size: 24px;
# # REMOVED_BY_DARK_THEME_PATCH:     font-weight: 700;
# # REMOVED_BY_DARK_THEME_PATCH:     letter-spacing: 1px;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: .erp-title span { color: #ef4444; }
# # REMOVED_BY_DARK_THEME_PATCH: .erp-sub {
# # REMOVED_BY_DARK_THEME_PATCH:     color: #94a3b8;
# # REMOVED_BY_DARK_THEME_PATCH:     font-size: 14px;
# # REMOVED_BY_DARK_THEME_PATCH:     margin-top: 4px;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* ===== CARDS ===== */
# # REMOVED_BY_DARK_THEME_PATCH: .card {
# # REMOVED_BY_DARK_THEME_PATCH:     background: white;
# # REMOVED_BY_DARK_THEME_PATCH:     padding: 20px;
# # REMOVED_BY_DARK_THEME_PATCH:     border-radius: 12px;
# # REMOVED_BY_DARK_THEME_PATCH:     border: 1px solid #e2e8f0;
# # REMOVED_BY_DARK_THEME_PATCH:     margin-bottom: 16px;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* ===== UPLOAD ===== */
# # REMOVED_BY_DARK_THEME_PATCH: [data-testid="stFileUploader"] {
# # REMOVED_BY_DARK_THEME_PATCH:     background-color: #f8fafc;
# # REMOVED_BY_DARK_THEME_PATCH:     border: 2px dashed #cbd5e1;
# # REMOVED_BY_DARK_THEME_PATCH:     border-radius: 12px;
# # REMOVED_BY_DARK_THEME_PATCH:     padding: 25px;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* ===== BOTÕES ===== */
# # REMOVED_BY_DARK_THEME_PATCH: div[data-testid="stButton"] > button[kind="primary"] {
# # REMOVED_BY_DARK_THEME_PATCH:     background-color: #2563eb !important;
# # REMOVED_BY_DARK_THEME_PATCH:     color: white !important;
# # REMOVED_BY_DARK_THEME_PATCH:     border-radius: 8px !important;
# # REMOVED_BY_DARK_THEME_PATCH:     font-weight: 700 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     font-size: 15px !important;
# # REMOVED_BY_DARK_THEME_PATCH:     padding: 12px 0 !important;
# # REMOVED_BY_DARK_THEME_PATCH:     border: none !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: div[data-testid="stButton"] > button[kind="primary"]:hover {
# # REMOVED_BY_DARK_THEME_PATCH:     background-color: #1d4ed8 !important;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* ===== MÉTRICAS ===== */
# # REMOVED_BY_DARK_THEME_PATCH: .metric-card {
# # REMOVED_BY_DARK_THEME_PATCH:     background: #f1f5f9;
# # REMOVED_BY_DARK_THEME_PATCH:     padding: 16px;
# # REMOVED_BY_DARK_THEME_PATCH:     border-radius: 10px;
# # REMOVED_BY_DARK_THEME_PATCH:     text-align: center;
# # REMOVED_BY_DARK_THEME_PATCH:     border: 1px solid #e2e8f0;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: .metric-card b {
# # REMOVED_BY_DARK_THEME_PATCH:     display: block;
# # REMOVED_BY_DARK_THEME_PATCH:     color: #64748b;
# # REMOVED_BY_DARK_THEME_PATCH:     font-size: 12px;
# # REMOVED_BY_DARK_THEME_PATCH:     font-weight: 600;
# # REMOVED_BY_DARK_THEME_PATCH:     text-transform: uppercase;
# # REMOVED_BY_DARK_THEME_PATCH:     letter-spacing: 0.5px;
# # REMOVED_BY_DARK_THEME_PATCH:     margin-bottom: 6px;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: .metric-card .metric-val {
# # REMOVED_BY_DARK_THEME_PATCH:     font-size: 22px;
# # REMOVED_BY_DARK_THEME_PATCH:     font-weight: 700;
# # REMOVED_BY_DARK_THEME_PATCH:     color: #0f172a;
# # REMOVED_BY_DARK_THEME_PATCH: }
# # REMOVED_BY_DARK_THEME_PATCH: .metric-card .metric-val.verde { color: #16a34a; }
# # REMOVED_BY_DARK_THEME_PATCH: .metric-card .metric-val.vermelho { color: #dc2626; }
# # REMOVED_BY_DARK_THEME_PATCH: .metric-card .metric-val.azul { color: #2563eb; }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* ===== ALERTS ===== */
# # REMOVED_BY_DARK_THEME_PATCH: .stAlert { border-radius: 10px; }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: /* ===== SCROLL ===== */
# # REMOVED_BY_DARK_THEME_PATCH: ::-webkit-scrollbar { width: 6px; }
# # REMOVED_BY_DARK_THEME_PATCH: ::-webkit-scrollbar-thumb { background: #64748b; border-radius: 10px; }
# # REMOVED_BY_DARK_THEME_PATCH: 
# # REMOVED_BY_DARK_THEME_PATCH: </style>
# # REMOVED_BY_DARK_THEME_PATCH: """, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# Estado da sessão
# ─────────────────────────────────────────────
if "lancamentos"         not in st.session_state: st.session_state.lancamentos         = []
if "historico"           not in st.session_state: st.session_state.historico           = []
if "log_gravacao"        not in st.session_state: st.session_state.log_gravacao        = []
if "erro_gravacao"       not in st.session_state: st.session_state.erro_gravacao       = ""
if "gravacao_confirmada" not in st.session_state: st.session_state.gravacao_confirmada = False
if "duplicatas"          not in st.session_state: st.session_state.duplicatas          = []


@st.dialog("Confirmar Gravação na PCLANC", width="large")
def _dialog_confirmacao():
    lancs       = st.session_state.lancamentos
    valor_total = sum(l["valor"] for l in lancs)
    duplicatas  = st.session_state.duplicatas

    st.markdown(
        f"Você está prestes a gravar **{len(lancs)} lançamentos** na PCLANC. "
        f"Revise o resumo abaixo antes de confirmar."
    )

    c1, c2, c3 = st.columns(3)
    c1.metric("Lançamentos", len(lancs))
    c2.metric("RCAs únicos", len({l["codusur"] for l in lancs}))
    c3.metric("Total", brl(valor_total))

    if duplicatas:
        nomes_dup = ", ".join(sorted({
            l["nome_rca"] if l["nome_rca"] else f"CODUSUR {l['codusur']}"
            for l in duplicatas
        }))
        st.warning(
            f"⚠ **{len(duplicatas)} lançamento(s) possivelmente duplicado(s)** já existem na PCLANC hoje "
            f"com os mesmos dados: {nomes_dup}. Confirme apenas se tiver certeza."
        )

    st.dataframe(
        pd.DataFrame([{
            "CODUSUR":   l["codusur"],
            "Nome RCA":  l["nome_rca"],
            "Parcela":   l["parcela"],
            "Valor":     brl(l["valor"]),
            "DTVENC":    l["dtvenc"].strftime("%d/%m/%Y") if l["dtvenc"] else "",
            "Histórico": l["historico"],
        } for l in lancs]),
        use_container_width=True,
        hide_index=True,
    )

    st.markdown("")
    btn1, btn2 = st.columns(2)
    if btn1.button("✔  Confirmar e Gravar", type="primary", use_container_width=True):
        st.session_state.gravacao_confirmada = True
        st.rerun()
    if btn2.button("✘  Cancelar", use_container_width=True):
        st.rerun()

# ─────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────
# # REMOVED_BY_BANNER_PATCH: st.sidebar.markdown("""
# # REMOVED_BY_BANNER_PATCH: <div style="padding:4px 0 8px 0;">
# # REMOVED_BY_BANNER_PATCH:     <div style="font-size:20px;font-weight:900;color:#f8fafc;letter-spacing:3px;">
# # REMOVED_BY_BANNER_PATCH:         ROFE <span style="color:#ef4444;">|</span>
# # REMOVED_BY_BANNER_PATCH:     </div>
# # REMOVED_BY_BANNER_PATCH:     <div style="font-size:11px;color:#94a3b8;letter-spacing:1.5px;margin-top:2px;">
# # REMOVED_BY_BANNER_PATCH:         COMISSÃO RCA
# # REMOVED_BY_BANNER_PATCH:     </div>
# # REMOVED_BY_BANNER_PATCH: </div>
# # REMOVED_BY_BANNER_PATCH: """, unsafe_allow_html=True)

# === SIDEBAR_ADMIN_PATCH ===
# Defaults pra quando usuario nao-admin (sidebar nao mostra os campos)
filial = CODFILIAL_PADRAO
conta = str(CODCONTA_PADRAO)
tipo_servico = TIPOSERVICO
tipo_lanc_op = "C"
parceiro_op = "R"
moeda_op = "R"
nf_op = "N"

if auth.is_admin():

    _slot_teste = st.sidebar.empty()
    if st.sidebar.button("🔌 Testar Conexão", use_container_width=True):
        ok, msg = testar_conexao()
        if ok:
            _slot_teste.success(msg)
        else:
            _slot_teste.error(msg)

    st.sidebar.markdown("---")
    st.sidebar.markdown('<p style="font-size:11px;font-weight:700;color:#64748b;letter-spacing:1px;margin:0 0 4px 0;">📌 IDENTIFICAÇÃO</p>', unsafe_allow_html=True)

    _c1, _c2 = st.sidebar.columns(2)
    filial = _c1.text_input("Filial", value=CODFILIAL_PADRAO, disabled=True, key="sb_filial")
    conta  = _c2.text_input("Conta",  value=str(CODCONTA_PADRAO), disabled=True, key="sb_conta")
    tipo_servico = st.sidebar.text_input("Tipo Serviço", value=TIPOSERVICO, disabled=True, key="sb_tpserv")

    st.sidebar.markdown("---")
    st.sidebar.markdown('<p style="font-size:11px;font-weight:700;color:#64748b;letter-spacing:1px;margin:0 0 4px 0;">💰 LANÇAMENTO</p>', unsafe_allow_html=True)

    _c3, _c4 = st.sidebar.columns(2)
    tipo_lanc_op = _c3.selectbox("Tipo Lanç.", ["C"], disabled=True, key="sb_tplanc")
    parceiro_op  = _c4.selectbox("Parceiro",   ["R"], disabled=True, key="sb_tpparc")

    _c5, _c6 = st.sidebar.columns(2)
    moeda_op = _c5.selectbox("Moeda",    ["R"], disabled=True, key="sb_moeda")
    nf_op    = _c6.text_input("NF Serv.", value="N", disabled=True, key="sb_nfserv")
# === FIM_SIDEBAR_ADMIN_PATCH ===

st.sidebar.markdown("---")
st.sidebar.download_button(
    label="⬇  Baixar Modelo Excel",
    data=gerar_modelo_excel(),
    file_name="modelo_comissao_rca.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

# Valores efetivos da sidebar (para o INSERT)
_filial_ativo = filial.strip() or CODFILIAL_PADRAO
try:
    _conta_ativa = int(conta.strip())
except ValueError:
    _conta_ativa = CODCONTA_PADRAO

# ─────────────────────────────────────────────
# Header
# ─────────────────────────────────────────────
# # REMOVED_BY_BANNER_PATCH: st.markdown("""
# # REMOVED_BY_BANNER_PATCH: <div class="erp-header">
# # REMOVED_BY_BANNER_PATCH:     <div class="erp-title">ROFE <span>|</span></div>
# # REMOVED_BY_BANNER_PATCH:     <div class="erp-sub">Importação de Comissões RCA — Rotina 749</div>
# # REMOVED_BY_BANNER_PATCH: </div>
# # REMOVED_BY_BANNER_PATCH: """, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# Cards de resumo
# ─────────────────────────────────────────────
_lanc_count  = len(st.session_state.lancamentos)
_ok_count    = sum(1 for l in st.session_state.log_gravacao if l.get("Status") == "OK")
_err_count   = sum(1 for l in st.session_state.log_gravacao if l.get("Status") != "OK") if st.session_state.log_gravacao else 0
_status_txt  = "Aguardando"
_status_cls  = ""
if st.session_state.lancamentos and not st.session_state.log_gravacao:
    _status_txt = "Pronto"; _status_cls = "azul"
elif st.session_state.log_gravacao:
    _status_txt = "Gravado" if _err_count == 0 else "Com erros"
    _status_cls = "verde"   if _err_count == 0 else "vermelho"

col1, col2, col3, col4 = st.columns(4)
col1.markdown(f'<div class="metric-card"><b>Registros carregados</b><div class="metric-val">{_lanc_count}</div></div>', unsafe_allow_html=True)
col2.markdown(f'<div class="metric-card"><b>Gravados com sucesso</b><div class="metric-val verde">{_ok_count}</div></div>', unsafe_allow_html=True)
col3.markdown(f'<div class="metric-card"><b>Erros</b><div class="metric-val vermelho">{_err_count}</div></div>', unsafe_allow_html=True)
col4.markdown(f'<div class="metric-card"><b>Status</b><div class="metric-val {_status_cls}">{_status_txt}</div></div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# Upload
# ─────────────────────────────────────────────
st.markdown("### 📂 Importação de Arquivo")

with st.container():
    st.markdown('<div class="card">', unsafe_allow_html=True)

    arquivo = st.file_uploader(
        "Arraste ou selecione o arquivo (.xlsx)",
        type=["xlsx"],
    )

    if arquivo:
        comissoes, erro = ler_excel(arquivo.read(), codfilial_padrao=_filial_ativo)

        if erro:
            st.error(f"Erro ao ler planilha: {erro}")
        else:
            nomes_rca   = {}   # cod -> nome
            bloqueios   = {}   # cod -> valor BLOQUEIO
            erros_rca   = []
            with st.spinner("Buscando nomes dos RCAs no Oracle..."):
                db = OracleConnection()
                if db.conectar():
                    try:
                        codigos = list({c.codusur for c in comissoes})
                        placeholders = ", ".join(f":cod{i}" for i in range(len(codigos)))
                        bind = {f"cod{i}": str(cod) for i, cod in enumerate(codigos)}
                        cur = db.get_cursor()
                        cur.execute(f"""
                            SELECT CODUSUR,
                                   NVL(NOME, '(sem nome)') AS NOME,
                                   NVL(BLOQUEIO, 'N')      AS BLOQUEIO
                              FROM PCUSUARI
                             WHERE CODUSUR IN ({placeholders})
                        """, bind)
                        for row in cur.fetchall():
                            cod = int(row[0])
                            nomes_rca[cod] = row[1]
                            bloqueios[cod] = str(row[2]).strip().upper()
                        cur.close()
                        for cod in codigos:
                            if cod not in nomes_rca:
                                nomes_rca[cod] = None
                                erros_rca.append(f"CODUSUR {cod} não encontrado")
                    except Exception as e:
                        erros_rca.append(f"Erro na consulta Oracle: {e}")
                    db.desconectar()
            # === AUDIT_PATCH === registra resultado
            if _imp_id:
                try:
                    audit.registrar_lancamentos(
                        importacao_id=_imp_id,
                        lancamentos=st.session_state.lancamentos,
                        recnums=recnums if sucesso else None,
                        sucesso=sucesso,
                        erro_msg=erro_lote if not sucesso else None,
                    )
                    audit.finalizar_importacao(
                        _imp_id,
                        sucesso=sucesso,
                        erro_msg=erro_lote if not sucesso else None,
                    )
                except Exception as _e:
                    print(f'[audit] falha registrar/finalizar: {_e}')
            # === FIM_AUDIT_POS ===
                else:
                    st.warning("Não foi possível conectar ao Oracle. Verifique o Instant Client.")

            # CODUSURs não encontrados no Oracle
            _nao_encontrados = {c.codusur for c in comissoes if nomes_rca.get(c.codusur) is None}

            lancamentos = []
            for c in comissoes:
                nome     = nomes_rca.get(c.codusur) or c.nome_rca or ""
                bloqueio = bloqueios.get(c.codusur, "")
                hist     = (c.historico or "").strip()
                # alerta = bloqueado, não encontrado ou historico vazio
                alerta   = bloqueio == "S" or c.codusur in _nao_encontrados or not hist
                lancamentos.append({
                    "linha":     c.linha,
                    "parcela":   c.parcela,
                    "codusur":   c.codusur,
                    "nome_rca":  nome,
                    "valor":     c.valor,
                    "codfilial": c.codfilial or _filial_ativo,
                    "codconta":  c.codconta  or _conta_ativa,
                    "dtvenc":    c.dtvenc,
                    "historico": hist or "COMISSAO RCA",
                    "bloqueio":  bloqueio,
                    "alerta":    alerta,
                })

            st.session_state.lancamentos   = lancamentos
            st.session_state.log_gravacao  = []
            st.session_state.erro_gravacao = ""

            if erros_rca:
                st.warning("RCAs não encontrados:\n" + "\n".join(f"• {e}" for e in erros_rca))

            # Badges
            rcas_unicos  = len({l["codusur"] for l in lancamentos})
            valor_tot    = sum(l["valor"]    for l in lancamentos)
            parcelas_n   = len({l["parcela"] for l in lancamentos})
            bloq_alerta  = [l for l in lancamentos if l["bloqueio"] == "S"]

            st.markdown(
                badge(f"✔ {len(lancamentos)} lançamentos", "verde")
                + badge(f"👤 {rcas_unicos} RCA(s)", "azul")
                + badge(f"📋 {parcelas_n} parcela(s)", "azul")
                + (badge(f"⚠ {len(erros_rca)} não encontrado(s)", "vermelho") if erros_rca else "")
                + (badge(f"⚠ {len({l['codusur'] for l in bloq_alerta})} RCA(s) BLOQUEADO(S)", "laranja") if bloq_alerta else ""),
                unsafe_allow_html=True,
            )
            st.markdown("")

            # Tabela — linha amarela quando alerta=True (não encontrado OU bloqueado)
            df_prev = pd.DataFrame([{
                "Linha":     l["linha"],
                "Parcela":   l["parcela"],
                "CODUSUR":   l["codusur"],
                "Nome RCA":  l["nome_rca"],
                "Valor":     brl(l["valor"]),
                "DTVENC":    l["dtvenc"].strftime("%d/%m/%Y") if l["dtvenc"] else "",
                "Histórico": l["historico"],
            } for l in lancamentos])

            _alert_flags = [l["alerta"] for l in lancamentos]

            def _highlight(row):
                cor = "background-color:#fef9c3; color:#78350f"
                return [cor] * len(row) if _alert_flags[row.name] else [""] * len(row)

            st.dataframe(
                df_prev.style.apply(_highlight, axis=1),
                use_container_width=True,
                hide_index=True,
            )

    # # REMOVED_BY_ORPHAN_DIV_PATCH: st.markdown('</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────
# Gravar
# ─────────────────────────────────────────────
if st.session_state.lancamentos:
    valor_total = sum(l["valor"] for l in st.session_state.lancamentos)
    st.markdown(
        badge(f"{len(st.session_state.lancamentos)} lançamentos prontos", "navy")
        + badge(f"Total: {brl(valor_total)}", "azul"),
        unsafe_allow_html=True,
    )
    st.markdown("")

    _alertados = [l for l in st.session_state.lancamentos if l.get("alerta")]
    if _alertados:
        _nomes_alerta = ", ".join(sorted({
            l["nome_rca"] if l["nome_rca"] else f"CODUSUR {l['codusur']}"
            for l in _alertados
        }))
        st.error(
            f"⛔ Gravação bloqueada — existem pendências nos RCAs: "
            f"**{_nomes_alerta}**. Corrija os alertas antes de prosseguir."
        )

    if st.button("▶  Executar Gravação na PCLANC", type="primary", use_container_width=True, disabled=bool(_alertados)):
        # Guarda dupla: impede gravação se houver alertas no momento do clique
        if [l for l in st.session_state.lancamentos if l.get("alerta")]:
            st.session_state.erro_gravacao = (
                "Gravação cancelada: existem RCAs com pendências. Corrija os alertas e tente novamente."
            )
        else:
            # Verificar duplicatas antes de abrir o dialog
            with st.spinner("Verificando duplicatas no Oracle..."):
                _db_dup = OracleConnection()
                if _db_dup.conectar():
                    st.session_state.duplicatas = GravadorPCLANC(_db_dup).verificar_duplicatas(
                        st.session_state.lancamentos
                    )
                    _db_dup.desconectar()
                else:
                    st.session_state.duplicatas = []
            _dialog_confirmacao()

    # Execução após confirmação no dialog
    if st.session_state.gravacao_confirmada:
        st.session_state.gravacao_confirmada = False
        st.session_state.duplicatas          = []

        with st.spinner("Conectando ao Oracle..."):
            db      = OracleConnection()
            ok_conn = db.conectar()

        if not ok_conn:
            st.session_state.erro_gravacao = "Não foi possível conectar ao Oracle."
        else:
            # === AUDIT_PATCH === inicia auditoria
            _imp_id = None
            try:
                _u = auth.usuario_logado()
                _arquivo_nome = st.session_state.get('arquivo_nome') or '(sem nome)'
                _total_itens = len(st.session_state.lancamentos)
                _valor_total = sum(float(_l['valor']) for _l in st.session_state.lancamentos)
                if _u:
                    _imp_id = audit.iniciar_importacao(
                        usuario_id=_u['id'],
                        arquivo_nome=_arquivo_nome,
                        total_itens=_total_itens,
                        valor_total=_valor_total,
                    )
            except Exception as _e:
                print(f'[audit] falha iniciar_importacao: {_e}')
            # === FIM_AUDIT_PRE ===
            gravador   = GravadorPCLANC(db)
            _prog_slot = st.empty()

            _prog_slot.progress(0.0, text="Gravando lote na PCLANC...")
            sucesso, recnums, erro_lote = gravador.inserir_lote(st.session_state.lancamentos)
            db.desconectar()

            if sucesso:
                _prog_slot.progress(1.0, text=f"Concluído — {len(recnums)} lançamento(s) gravado(s)")
                log_gravacao = [
                    {
                        "RECNUM":    recnum,
                        "CODUSUR":   r["codusur"],
                        "Nome RCA":  r["nome_rca"],
                        "Parcela":   r["parcela"],
                        "Valor":     brl(r["valor"]),
                        "DTVENC":    r["dtvenc"].strftime("%d/%m/%Y") if r["dtvenc"] else "",
                        "Histórico": r["historico"],
                        "Status":    "OK",
                    }
                    for recnum, r in zip(recnums, st.session_state.lancamentos)
                ]
            else:
                _prog_slot.progress(1.0, text="Erro — rollback realizado")
                log_gravacao = [
                    {
                        "RECNUM":    "-",
                        "CODUSUR":   r["codusur"],
                        "Nome RCA":  r["nome_rca"],
                        "Parcela":   r["parcela"],
                        "Valor":     brl(r["valor"]),
                        "DTVENC":    r["dtvenc"].strftime("%d/%m/%Y") if r["dtvenc"] else "",
                        "Histórico": r["historico"],
                        "Status":    f"ERRO: {erro_lote}",
                    }
                    for r in st.session_state.lancamentos
                ]

            st.session_state.log_gravacao  = log_gravacao
            st.session_state.erro_gravacao = "" if sucesso else f"Erro na gravação — rollback realizado: {erro_lote}"

            from datetime import datetime as _dt
            _agora = _dt.now().strftime("%d/%m/%Y %H:%M:%S")
            for item in log_gravacao:
                st.session_state.historico.append({**item, "Gravado em": _agora})

# Resultado — fora do if button para persistir
if st.session_state.erro_gravacao:
    st.error(st.session_state.erro_gravacao)

if st.session_state.log_gravacao:
    log       = st.session_state.log_gravacao
    ok_count  = sum(1 for l in log if l["Status"] == "OK")
    err_count = len(log) - ok_count

    st.markdown(
        badge(f"✔ {ok_count} gravado(s) com sucesso", "verde")
        + (badge(f"✘ {err_count} erro(s)", "vermelho") if err_count else ""),
        unsafe_allow_html=True,
    )
    st.markdown("")

    if err_count == 0:
        st.success(f"{ok_count} lançamentos gravados com sucesso na PCLANC!")
    else:
        st.error(f"{err_count} erro(s) durante a gravação.")

    st.markdown("**RECNUMs gerados**")
    st.dataframe(pd.DataFrame(log), use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────
# Histórico da Sessão
# ─────────────────────────────────────────────
st.markdown("### 📜 Histórico da Sessão")

with st.container():
    st.markdown('<div class="card">', unsafe_allow_html=True)

    if not st.session_state.historico:
        st.info("Nenhum lançamento realizado ainda.")
    else:
        df_hist = pd.DataFrame(st.session_state.historico)
        ok_df   = df_hist[df_hist["Status"] == "OK"]

        try:
            vth = (
                ok_df["Valor"]
                .str.replace("R\\$ ", "", regex=True)
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False)
                .astype(float).sum()
            )
            _vth_fmt = brl(vth)
        except Exception:
            _vth_fmt = "-"

        h1, h2, h3, h4 = st.columns(4)
        h1.markdown(f'<div class="metric-card"><b>Total gravado</b><div class="metric-val verde">{len(ok_df)}</div></div>', unsafe_allow_html=True)
        h2.markdown(f'<div class="metric-card"><b>Erros</b><div class="metric-val vermelho">{len(df_hist) - len(ok_df)}</div></div>', unsafe_allow_html=True)
        h3.markdown(f'<div class="metric-card"><b>RCAs únicos</b><div class="metric-val">{df_hist["CODUSUR"].nunique()}</div></div>', unsafe_allow_html=True)
        h4.markdown(f'<div class="metric-card"><b>Valor total gravado</b><div class="metric-val azul">{_vth_fmt}</div></div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.dataframe(df_hist, use_container_width=True, hide_index=True)

        def _hist_excel(df: pd.DataFrame) -> bytes:
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Historico")
                ws = writer.sheets["Historico"]
                for col in ws.columns:
                    max_len = max(len(str(c.value or "")) for c in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
            return buf.getvalue()

        c1, c2 = st.columns([2, 1])
        with c1:
            st.download_button(
                label="⬇  Baixar Relatório (Excel)",
                data=_hist_excel(df_hist),
                file_name="historico_comissoes_rca.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with c2:
            if st.button("🗑 Limpar histórico", use_container_width=True):
                st.session_state.historico = []
                st.rerun()

    # # REMOVED_BY_ORPHAN_DIV_PATCH: st.markdown('</div>', unsafe_allow_html=True)
