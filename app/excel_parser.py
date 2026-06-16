"""
Módulo de leitura e parse da planilha Excel de comissões RCA.

Suporta dois formatos de planilha (auto-detectados):

  FORMATO A — modelo simples (gerado pelo botão "Baixar Modelo" do app):
      Aba única, cabeçalho na linha 1, com colunas:
          - parceiro(COD)  → CODUSUR do RCA
          - VALOR          → valor total da comissão

  FORMATO B — planilha gerencial completa que a operação já usa:
      Múltiplas abas (ex: "RCA'S GERAIS", "RCAS DIFERENCIADOS").
      Cabeçalho na linha 2 (linha 1 pode ter totais residuais).
      Colunas relevantes:
          - COD       → CODUSUR
          - A Pagar   → valor (em qualquer coluna)
      Demais colunas (Faturamento, Comissao, etc) são ignoradas.

Em ambos os formatos, a data de vencimento é informada no APP (não na planilha).

Regra de parcelamento:
  - valor <= LIMITE_PARCELA_UNICA  → 1 lançamento  com dt_parcela_1
  - valor >  LIMITE_PARCELA_UNICA  → 2 lançamentos (valor/2 cada)
                                     parcela 1: round half up    com dt_parcela_1
                                     parcela 2: total - parcela1 com dt_parcela_2
"""
from __future__ import annotations

import openpyxl
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import List, Optional, Tuple

from models import ComissaoRCA
from config import CODCONTA_PADRAO, CODFILIAL_PADRAO


# Limite (R$) abaixo do qual o lançamento NÃO é parcelado.
LIMITE_PARCELA_UNICA = 2000.00

# Quantas linhas no início da planilha procurar pelo cabeçalho.
_LINHAS_PROCURAR_CABECALHO = 5


def _meio_arredondar_cima(valor_total: float) -> float:
    """Retorna valor_total/2 com ROUND_HALF_UP na 2ª casa decimal.

    Exemplo:
      2.219,49 / 2 = 1.109,745 → 1.109,75
      2.219,48 / 2 = 1.109,74  → 1.109,74
    """
    metade = Decimal(str(valor_total)) / Decimal(2)
    return float(metade.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


# =============================================================================
# Entrypoint
# =============================================================================
def ler_excel(
    conteudo_bytes: bytes,
    dt_parcela_1: date,
    dt_parcela_2: date,
    historico: str = "COMISSAO RCA",
    codfilial_padrao: str = CODFILIAL_PADRAO,
    limite_parcela_unica: float = LIMITE_PARCELA_UNICA,
) -> Tuple[List[ComissaoRCA], Optional[str]]:
    """Lê a planilha e retorna lista de ComissaoRCA aplicando a regra de parcelamento.

    Lê de TODAS as abas que tiverem cabeçalho compatível.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(conteudo_bytes), read_only=True, data_only=True)

        comissoes: List[ComissaoRCA] = []
        avisos: List[str] = []
        sheets_lidas: List[str] = []
        sheets_puladas: List[str] = []

        for ws in wb.worksheets:
            sheet_comissoes, motivo = _ler_aba(
                ws,
                dt_parcela_1=dt_parcela_1,
                dt_parcela_2=dt_parcela_2,
                historico=historico,
                codfilial_padrao=codfilial_padrao,
                limite_parcela_unica=limite_parcela_unica,
            )
            if sheet_comissoes:
                comissoes.extend(sheet_comissoes)
                sheets_lidas.append(f"{ws.title} ({len(sheet_comissoes)} lançamentos)")
            else:
                sheets_puladas.append(f"{ws.title}: {motivo}")

        wb.close()

        if not comissoes:
            msg = "Nenhum lançamento encontrado em nenhuma aba."
            if sheets_puladas:
                msg += " Detalhes:\n  - " + "\n  - ".join(sheets_puladas)
            return [], msg

        # Loga (em prod, isto vai pro stdout do container)
        print(f"[excel_parser] abas lidas: {', '.join(sheets_lidas)}")
        if sheets_puladas:
            print(f"[excel_parser] abas puladas: {', '.join(sheets_puladas)}")

        return comissoes, None

    except Exception as e:
        return [], f"Erro ao ler planilha: {e}"


# =============================================================================
# Lê uma aba específica
# =============================================================================
def _ler_aba(
    ws,
    dt_parcela_1: date,
    dt_parcela_2: date,
    historico: str,
    codfilial_padrao: str,
    limite_parcela_unica: float,
) -> Tuple[List[ComissaoRCA], Optional[str]]:
    """Lê uma aba e retorna (lista, motivo_de_falha_se_houver).

    Procura o cabeçalho nas primeiras N linhas — uma linha conta como
    cabeçalho quando contém PELO MENOS as colunas codusur+valor.
    """
    # Pega as primeiras N linhas
    primeiras = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=_LINHAS_PROCURAR_CABECALHO,
                                          values_only=True), start=1):
        primeiras.append((i, row))

    if not primeiras:
        return [], "aba vazia"

    # Procura a linha do cabeçalho
    header_row_num = None
    col_map = None
    for linha_num, row in primeiras:
        if row is None:
            continue
        header_candidate = [
            str(v).strip() if v is not None else ""
            for v in row
        ]
        cm = _mapear_colunas(header_candidate)
        if "codusur" in cm and "valor" in cm:
            header_row_num = linha_num
            col_map = cm
            break

    if header_row_num is None:
        # Pra ajudar o diagnóstico, mostra o que viu na primeira linha não-vazia
        for linha_num, row in primeiras:
            if row and any(v is not None for v in row):
                vista = [str(v).strip() if v else "" for v in row[:10]]
                return [], (
                    f"cabeçalho não encontrado nas primeiras {_LINHAS_PROCURAR_CABECALHO} linhas "
                    f"(esperando COD/CODUSUR + A Pagar/VALOR). Linha {linha_num} mostra: "
                    + ", ".join(v for v in vista if v)
                )
        return [], "aba sem dados"

    # Parse dos dados a partir da linha seguinte ao cabeçalho
    comissoes: List[ComissaoRCA] = []
    for row_num, row in enumerate(
        ws.iter_rows(min_row=header_row_num + 1, values_only=True),
        start=header_row_num + 1,
    ):
        if row is None or all(v is None for v in row):
            continue

        codusur_val = _get_cell(row, col_map.get("codusur"))
        if codusur_val is None:
            continue
        try:
            codusur = int(float(str(codusur_val)))
        except (ValueError, TypeError):
            continue
        if codusur <= 0:
            continue

        valor_val = _get_cell(row, col_map.get("valor"))
        if valor_val is None:
            continue
        try:
            valor_total = _parse_valor(valor_val)
        except (ValueError, TypeError):
            continue
        if valor_total <= 0:
            continue

        nome_rca  = _parse_str(_get_cell(row, col_map.get("nome_rca"))) or ""
        codconta  = _parse_int(_get_cell(row, col_map.get("codconta"))) or CODCONTA_PADRAO
        codfilial = _parse_str(_get_cell(row, col_map.get("codfilial"))) or codfilial_padrao
        historico_linha = historico

        # ── Aplica regra de parcelamento ────────────────────────────────
        if valor_total <= limite_parcela_unica:
            comissoes.append(ComissaoRCA(
                linha=row_num,
                parcela=1,
                codusur=codusur,
                nome_rca=nome_rca,
                valor=round(valor_total, 2),
                codconta=codconta,
                codfilial=codfilial,
                historico=historico_linha,
                dtvenc=dt_parcela_1,
            ))
        else:
            valor_p1 = _meio_arredondar_cima(valor_total)
            valor_p2 = round(valor_total - valor_p1, 2)
            comissoes.append(ComissaoRCA(
                linha=row_num,
                parcela=1,
                codusur=codusur,
                nome_rca=nome_rca,
                valor=valor_p1,
                codconta=codconta,
                codfilial=codfilial,
                historico=historico_linha,
                dtvenc=dt_parcela_1,
            ))
            comissoes.append(ComissaoRCA(
                linha=row_num,
                parcela=2,
                codusur=codusur,
                nome_rca=nome_rca,
                valor=valor_p2,
                codconta=codconta,
                codfilial=codfilial,
                historico=historico_linha,
                dtvenc=dt_parcela_2,
            ))

    if not comissoes:
        return [], f"cabeçalho na linha {header_row_num} mas sem dados válidos"
    return comissoes, None


# =============================================================================
# Aliases de coluna
# =============================================================================
_ALIASES = {
    "codusur":   ["COD", "CODUSUR", "COD_USUR", "PARCEIRO(COD)", "PARCEIRO",
                  "CODIGO_RCA", "COD_RCA", "RCA_COD", "CODRCA"],
    "nome_rca":  ["RCA", "NOME_RCA", "NOME RCA", "NOME", "PARCEIRO_NOME"],
    "codconta":  ["CONTADEBITO", "CONTA DEBITO", "CONTA_DEBITO",
                  "CODCONTA", "COD_CONTA", "CONTA", "CODIGO_CONTA"],
    "codfilial": ["CODFILIAL", "COD_FILIAL", "FILIAL", "CODIGO_FILIAL"],
    "historico": ["HISTORICO", "HISTÓRICO", "DESCRICAO", "DESCRIÇÃO", "OBS"],
    "valor":     ["A PAGAR", "APAGAR", "A_PAGAR",
                  "VALOR", "VALOR_TOTAL", "VALORTOTAL", "VALOR TOTAL",
                  "TOTAL", "VALOR_COMISSAO", "COMISSAO"],
}


def _mapear_colunas(header: List[str]) -> dict:
    col_map: dict = {}
    for idx, col_name in enumerate(header):
        if not col_name:
            continue
        col_upper = col_name.strip().upper()
        # normaliza espaços múltiplos
        col_upper_norm = " ".join(col_upper.split())
        for campo, nomes in _ALIASES.items():
            if campo in col_map:
                continue
            for n in nomes:
                if col_upper_norm == n or col_upper == n:
                    col_map[campo] = idx
                    break
    return col_map


# =============================================================================
# Helpers
# =============================================================================
def _get_cell(row: tuple, idx: Optional[int]):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_valor(val) -> float:
    """Converte valor monetário para float (aceita R$, ponto de milhar, vírgula decimal)."""
    if isinstance(val, (int, float)):
        return float(val)
    texto = str(val).strip()
    texto = texto.replace("R$", "").replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    return float(texto)


def _parse_int(val) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def _parse_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.upper() not in ("NONE", "NAN") else None
